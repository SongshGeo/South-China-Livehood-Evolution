#!/usr/bin/env python
"""Build the manuscript table workbook, ``paper/figs/SCE_Tables.xlsx``.

One sheet per table, named for the table's number; every sheet is row 1 caption,
row 2 header, data below, so each ``xlsx-table`` block carries ``skip_n: 1``.
The ``_index`` sheet registers where each sheet's numbers came from.

Values are read from their sources rather than retyped: parameter defaults from
the composed Hydra config, landscape counts from the input rasters, swept ranges
from ``run_slurm_rerun.sh``. Formulas are never written — ``xlsx_table.lua`` reads
cached cell values, and a formula written here would render blank in the PDF.

Run from the repository root::

    uv run python paper/build_tables.py
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import numpy as np
import rasterio
from hydra import compose, initialize_config_dir
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "paper" / "figs" / "SCE_Tables.xlsx"
BUILT = dt.date.today().isoformat()


# --------------------------------------------------------------------------
# Sources
# --------------------------------------------------------------------------
def load_config():
    """Effective baseline parameters, as Hydra actually composes them."""
    with initialize_config_dir(version_base=None, config_dir=str(ROOT / "config")):
        return compose(config_name="config")


def landscape_counts() -> dict[str, int | float]:
    """Cell counts and mean cell area, recomputed from the input rasters.

    Mirrors ``Env.setup_dem`` and ``CompetingCell``: the grid is the DEM's valid
    cells, negative slope readings are masked exactly as ``_open_rasterio`` does,
    and the arability predicates use the thresholds in ``src/api/env.py``.
    """
    with rasterio.open(ROOT / "data" / "ohndem10.tif") as d:
        dem = d.read(1).astype(float)
        res, bounds, shape = d.res[0], d.bounds, d.shape
    with rasterio.open(ROOT / "data" / "ohnslo10.tif") as d:
        slope = d.read(1).astype(float)
    with rasterio.open(ROOT / "data" / "ohn_waterbody.tif") as d:
        water = d.read(1).astype(float)

    grid = dem > -1000  # ABSESpy drops the DEM's nodata from the grid
    dem = np.where(dem < 0, np.nan, dem)
    slope = np.where(slope < 0, np.nan, slope)

    with np.errstate(invalid="ignore"):
        lowland = (dem > 0) & (dem < 200)
        arable = grid & lowland & (slope <= 10)
        rice = grid & lowland & (slope <= 0.5)
    near_water = grid & (water == 1)

    # Mean cell area on a spherical Earth, averaged over the **modelled** cells, not
    # over the whole raster frame. The frame includes the nodata rows that masking
    # removes, and they sit at latitudes the model never occupies: averaging over the
    # frame gives 78.23 km² where the 6835 modelled cells give 78.30. The SI and
    # finding F8 both quote the modelled figure -- F8's 80 km²/cell divisor is rounded
    # from it -- so that is the one this file must produce.
    radius_km = 6371.0088
    lats = np.array([bounds.top - (i + 0.5) * res for i in range(shape[0])])
    dlat = res * np.pi / 180 * radius_km
    dlon = dlat * np.cos(np.radians(lats))
    cell_area = (dlat * dlon)[:, None] * np.ones((1, shape[1]))
    area = float(cell_area[grid].mean())

    return {
        "cells": int(grid.sum()),
        "arable": int(arable.sum()),
        "rice_arable": int(rice.sum()),
        "near_water": int(near_water.sum()),
        "inland": int(grid.sum()) - int(near_water.sum()),
        "cell_area_km2": round(area, 1),
        # Two decimals, not one: the east edge is 121.2499...°, so one decimal forces a
        # choice between understating the coverage (121.2) and overstating it (121.3).
        "west": round(bounds.left, 2),
        "east": round(bounds.right, 2),
        "south": round(bounds.bottom, 2),
        "north": round(bounds.top, 2),
    }


def limh_swept() -> list[int]:
    """The forager-capacity arms actually dispatched, read from the SLURM script.

    Parsed from the `for lh in ...` loop in `run_slurm_rerun.sh` rather than
    written out here, for the same reason `swept_ranges` reads `F2H_VALUES`:
    a literal in this file drifts silently the next time the sweep changes, and
    the table goes on reporting a range no run ever used.
    `tests/test_sweep_scripts.py` locks the baseline to one of these arms.
    """
    text = (ROOT / "run_slurm_rerun.sh").read_text(encoding="utf-8")
    line = next(ln for ln in text.splitlines() if ln.strip().startswith("for lh in "))
    body = line.split("for lh in ", 1)[1].split(";", 1)[0]
    return [int(v) for v in body.split()]


def swept_ranges() -> dict[str, str]:
    """The conversion grid actually dispatched, read from the SLURM script.

    Read from the script that produced the published data, for the same reason
    `limh_swept` is: a literal here, or a read from a superseded script, goes on
    reporting a range no run used and nothing says so.
    """
    text = (ROOT / "run_slurm_rerun.sh").read_text(encoding="utf-8")
    out = {}
    for key, var in (("f2h", "F2H_VALUES"), ("h2f", "H2F_VALUES")):
        line = next(ln for ln in text.splitlines() if ln.startswith(f"{var}="))
        vals = [float(v) for v in line.split("(", 1)[1].rstrip(")").split()]
        out[key] = f"{vals[0]:g}–{vals[-1]:g} (step {vals[1] - vals[0]:g})"
    return out


# --------------------------------------------------------------------------
# Sheets
# --------------------------------------------------------------------------
def sheet(wb: Workbook, name: str, caption: str, header: list, rows: list[list]):
    ws = wb.create_sheet(name)
    ws.append([caption])
    ws.append(header)
    for r in rows:
        ws.append(r)
    for i, _ in enumerate(header, start=1):
        longest = max(
            [len(str(header[i - 1]))]
            + [len(str(r[i - 1])) for r in rows if i - 1 < len(r)]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 62)
    return ws


def build() -> None:
    cfg = load_config()
    land = landscape_counts()
    swept = swept_ranges()
    F, R, H = cfg.Farmer, cfg.RiceFarmer, cfg.Hunter
    env = cfg.env
    ceiling = int(env.lim_h * land["cells"])

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Table 1 (main text) ---------------------------------------------
    sheet(
        wb,
        "Table 1",
        "Baseline values and swept ranges for the parameters that govern how fast "
        "farming spreads; the two conversion probabilities are the only ones swept "
        "at fine resolution.",
        ["Symbol", "Quantity", "Unit", "Default", "Range tested"],
        [
            ["$T$", "Run length", "steps", cfg.time.end, "—"],
            ["$A$", "Modelled land cells", "cells", land["cells"], "—"],
            ["—", "Cell area", "km$^2$", land["cell_area_km2"], "—"],
            ["$r_F$", "Growth rate, rainfed farmer", "per step", F.growth_rate, "—"],
            ["$r_R$", "Growth rate, paddy farmer", "per step", R.growth_rate, "—"],
            ["$r_H$", "Growth rate, forager", "per step", H.growth_rate, "—"],
            [
                "$\\lambda_F$",
                "Immigration intensity, rainfed",
                "groups per step",
                env.lam_farmer,
                "2–10",
            ],
            [
                "$\\lambda_R$",
                "Immigration intensity, paddy",
                "groups per step",
                env.lam_ricefarmer,
                "0.1–0.5",
            ],
            [
                "$p_{F \\to H}$",
                "Conversion, rainfed to forager (f2h)",
                "per step",
                F.convert_prob.to_hunter,
                swept["f2h"],
            ],
            [
                "$p_{H \\to F}$",
                "Conversion, forager to rainfed (h2f)",
                "per step",
                H.convert_prob.to_farmer,
                swept["h2f"],
            ],
            [
                "$k_H$",
                "Forager carrying capacity",
                "people per cell",
                env.lim_h,
                ", ".join(str(v) for v in limh_swept()),
            ],
            ["$K_H$", "Regional forager ceiling", "people", ceiling, "—"],
            ["$c_F$", "Land per person, rainfed", "km$^2$", F.capital_area, "—"],
            ["$c_R$", "Land per person, paddy", "km$^2$", R.capital_area, "—"],
            ["$\\kappa$", "Intensification factor", "—", F.complexity, "—"],
            ["$N_{\\min}$", "Minimum viable group size", "people", F.min_size, "—"],
            [
                "—",
                "Initial forager cover",
                "fraction of land cells",
                env.init_hunters,
                "—",
            ],
            ["—", "Initial farming groups", "groups", env.init_farmers, "—"],
        ],
    )

    # ---- Table S1 — stochasticity ----------------------------------------
    sheet(
        wb,
        "Table S1",
        "Every source of randomness in the model, with the reason each is random "
        "rather than fixed.",
        ["Source", "Why it is random", "Implementation"],
        [
            [
                "Immigrant counts",
                "Arrivals from outside the region are unmodelled, so their timing is a rate rather than a schedule",
                "$\\mathrm{Poisson}(\\lambda_F)$, $\\mathrm{Poisson}(\\lambda_R)$",
            ],
            [
                "Immigrant and colony destination",
                "The model has no theory of site preference within the eligible set",
                "Uniform over eligible cells",
            ],
            [
                "Initial forager placement",
                "No data fix which cells were occupied at the start",
                "Uniform over land cells",
            ],
            [
                "Initial and colony group sizes",
                "The sources give ranges, not point values",
                "Uniform within bounds",
            ],
            [
                "Conversion trials",
                "Changing livelihood is treated as a propensity, not a determinate response",
                "Bernoulli at convert_prob",
            ],
            [
                "Colonization trial",
                "As above, for the decision to send out a colony",
                "Bernoulli at diffuse_prob",
            ],
            [
                "Mortality trial",
                "Shocks such as crop failure and disease are exogenous to the model",
                "Bernoulli at loss.prob",
            ],
            [
                "Forager movement target",
                "No theory of directional preference",
                "Uniform over the nearest eligible ring",
            ],
            [
                "Activation order",
                "Avoids an artefact from a fixed order under asynchronous updating",
                "Reshuffled each step",
            ],
            [
                "Trimming order at the ceiling",
                "No theory of which groups absorb the shortfall",
                "Reshuffled each step",
            ],
        ],
    )

    # ---- Table S2 — conversion paths --------------------------------------
    sheet(
        wb,
        "Table S2",
        "The five directed conversion paths, with the conditions and probability "
        "gating each; only the two paths linking rainfed farmers and foragers are "
        "swept.",
        ["Path", "Conditions", "Parameter", "Default", "Range tested"],
        [
            [
                "Farmer $\\to$ Hunter (f2h)",
                f"size $\\le$ {F.convert_threshold.to_hunter}",
                "Farmer.convert_prob.to_hunter",
                F.convert_prob.to_hunter,
                swept["f2h"],
            ],
            [
                "Hunter $\\to$ Farmer (h2f)",
                "A Farmer neighbour; cell is rainfed-arable",
                "Hunter.convert_prob.to_farmer",
                H.convert_prob.to_farmer,
                swept["h2f"],
            ],
            [
                "Farmer $\\to$ RiceFarmer",
                f"size $\\ge$ {F.convert_threshold.to_rice}; cell is paddy-arable",
                "Farmer.convert_prob.to_rice",
                F.convert_prob.to_rice,
                "—",
            ],
            [
                "Hunter $\\to$ RiceFarmer",
                "A RiceFarmer neighbour; cell is paddy-arable",
                "Hunter.convert_prob.to_rice",
                H.convert_prob.to_rice,
                "—",
            ],
            [
                "RiceFarmer $\\to$ Farmer",
                f"size $<$ {R.convert_threshold.to_farmer}",
                "RiceFarmer.convert_prob.to_farmer",
                R.convert_prob.to_farmer,
                "—",
            ],
        ],
    )

    # ---- Table S3 — symbol map -------------------------------------------
    sheet(
        wb,
        "Table S3",
        "One-to-one mapping between the main-text symbols and the implementation "
        "parameter names, so that the two descriptions of the model cannot drift "
        "apart.",
        ["Symbol", "Parameter", "Default", "Meaning"],
        [
            ["$T$", "time.end", cfg.time.end, "Steps per run"],
            ["$A$", "derived", land["cells"], "Modelled land cells"],
            ["$N$", "size", "—", "Group population"],
            ["$N_{\\min}$", "min_size", F.min_size, "Minimum viable group size"],
            [
                "$N_{\\max}$",
                "max_size",
                "3142 / 6283",
                "Per-group capacity, rainfed / paddy",
            ],
            ["$r_F$", "Farmer.growth_rate", F.growth_rate, "Rainfed growth rate"],
            ["$r_R$", "RiceFarmer.growth_rate", R.growth_rate, "Paddy growth rate"],
            ["$r_H$", "Hunter.growth_rate", H.growth_rate, "Forager growth rate"],
            [
                "$\\lambda_F$",
                "env.lam_farmer",
                env.lam_farmer,
                "Rainfed immigration intensity",
            ],
            [
                "$\\lambda_R$",
                "env.lam_ricefarmer",
                env.lam_ricefarmer,
                "Paddy immigration intensity",
            ],
            [
                "$p_{F \\to H}$",
                "Farmer.convert_prob.to_hunter",
                F.convert_prob.to_hunter,
                "Rainfed to forager",
            ],
            [
                "$p_{H \\to F}$",
                "Hunter.convert_prob.to_farmer",
                H.convert_prob.to_farmer,
                "Forager to rainfed",
            ],
            ["$k_H$", "env.lim_h", env.lim_h, "Forager capacity per cell"],
            ["$K_H$", "global_hunter_limit", ceiling, "Regional forager ceiling"],
            ["$a$", "area", F.area, "Cultivated radius (km)"],
            [
                "$c$",
                "capital_area",
                f"{F.capital_area} / {R.capital_area}",
                "Land per person, rainfed / paddy (km$^2$)",
            ],
            ["$\\kappa$", "complexity", F.complexity, "Intensification factor"],
            ["$s_t$", "derived", "—", "Agricultural share of total population"],
        ],
    )

    # ---- Table S4 — full parameter list -----------------------------------
    lit = "Literature"
    exp = "Expert judgement"
    der = "Derived"
    sheet(
        wb,
        "Table S4",
        "Every model parameter with its baseline value, the range over which it was "
        "swept, and whether the value is taken from the literature or set by expert "
        "judgement.",
        ["Parameter", "Meaning", "Default", "Range tested", "Provenance", "Source"],
        [
            ["time.end", "Steps per run", cfg.time.end, "—", exp, "—"],
            [
                "exp.repeats",
                "Replicates per combination",
                cfg.exp.repeats,
                "—",
                exp,
                "—",
            ],
            [
                "min_size",
                "Minimum viable group size",
                F.min_size,
                "—",
                lit,
                "Binford 2001; Kelly 2013",
            ],
            [
                "Farmer.growth_rate",
                "Growth rate, rainfed",
                F.growth_rate,
                "—",
                exp,
                "—",
            ],
            [
                "RiceFarmer.growth_rate",
                "Growth rate, paddy",
                R.growth_rate,
                "—",
                exp,
                "—",
            ],
            [
                "Hunter.growth_rate",
                "Growth rate, forager",
                H.growth_rate,
                "—",
                exp,
                "—",
            ],
            [
                "Farmer.area",
                "Cultivated radius (km)",
                F.area,
                "—",
                lit,
                "Shelach 1999; Wu et al. 2023",
            ],
            [
                "Farmer.capital_area",
                "Land per person, rainfed (km$^2$)",
                F.capital_area,
                "—",
                lit,
                "Qiao 2010, adjusted",
            ],
            [
                "RiceFarmer.capital_area",
                "Land per person, paddy (km$^2$)",
                R.capital_area,
                "—",
                exp,
                "—",
            ],
            [
                "Farmer.max_size",
                "Per-group capacity, rainfed",
                3142,
                "—",
                der,
                "$\\pi a^2 / c_F$",
            ],
            [
                "RiceFarmer.max_size",
                "Per-group capacity, paddy",
                6283,
                "—",
                der,
                "$\\pi a^2 / c_R$",
            ],
            ["complexity", "Intensification factor", F.complexity, "—", exp, "—"],
            [
                "env.lam_farmer",
                "Immigration intensity, rainfed",
                env.lam_farmer,
                "2–10",
                exp,
                "—",
            ],
            [
                "env.lam_ricefarmer",
                "Immigration intensity, paddy",
                env.lam_ricefarmer,
                "0.1–0.5",
                exp,
                "—",
            ],
            [
                "env.tick_farmer",
                "First step rainfed immigration may occur",
                env.tick_farmer,
                "—",
                exp,
                "—",
            ],
            [
                "env.tick_ricefarmer",
                "First step paddy immigration may occur",
                env.tick_ricefarmer,
                "—",
                exp,
                "—",
            ],
            [
                "Farmer.init_size",
                "Initial group size, rainfed",
                f"{F.init_size[0]}–{F.init_size[1]}",
                "—",
                exp,
                "Unused at baseline",
            ],
            [
                "RiceFarmer.init_size",
                "Initial group size, paddy",
                f"{R.init_size[0]}–{R.init_size[1]}",
                "—",
                exp,
                "Unused at baseline",
            ],
            [
                "Hunter.init_size",
                "Initial group size, forager",
                f"{H.init_size[0]}–{H.init_size[1]}",
                "—",
                exp,
                "—",
            ],
            [
                "Farmer.new_group_size",
                "Immigrant and colony size, rainfed",
                f"{F.new_group_size[0]}–{F.new_group_size[1]}",
                "—",
                exp,
                "—",
            ],
            [
                "RiceFarmer.new_group_size",
                "Immigrant and colony size, paddy",
                f"{R.new_group_size[0]}–{R.new_group_size[1]}",
                "—",
                exp,
                "—",
            ],
            [
                "Hunter.new_group_size",
                "Colony size, forager",
                f"{H.new_group_size[0]}–{H.new_group_size[1]}",
                "—",
                exp,
                "—",
            ],
            ["diffuse_prob", "Colonization probability", F.diffuse_prob, "—", exp, "—"],
            [
                "max_travel_distance",
                "Colonization and movement range (cells)",
                F.max_travel_distance,
                "—",
                exp,
                "—",
            ],
            [
                "Farmer.convert_prob.to_hunter",
                "Rainfed to forager (f2h)",
                F.convert_prob.to_hunter,
                swept["f2h"],
                exp,
                "—",
            ],
            [
                "Hunter.convert_prob.to_farmer",
                "Forager to rainfed (h2f)",
                H.convert_prob.to_farmer,
                swept["h2f"],
                exp,
                "—",
            ],
            [
                "Farmer.convert_prob.to_rice",
                "Rainfed to paddy",
                F.convert_prob.to_rice,
                "—",
                exp,
                "—",
            ],
            [
                "Hunter.convert_prob.to_rice",
                "Forager to paddy",
                H.convert_prob.to_rice,
                "—",
                exp,
                "—",
            ],
            [
                "RiceFarmer.convert_prob.to_farmer",
                "Paddy to rainfed",
                R.convert_prob.to_farmer,
                "—",
                exp,
                "—",
            ],
            [
                "convert_threshold.to_hunter",
                "Size below which rainfed may revert",
                F.convert_threshold.to_hunter,
                "—",
                exp,
                "—",
            ],
            [
                "convert_threshold.to_rice",
                "Size above which rainfed may take up paddy",
                F.convert_threshold.to_rice,
                "—",
                exp,
                "—",
            ],
            [
                "convert_threshold.to_farmer",
                "Size below which paddy reverts",
                R.convert_threshold.to_farmer,
                "—",
                exp,
                "—",
            ],
            [
                "Hunter.max_size",
                "Per-group forager capacity, inland",
                H.max_size,
                "—",
                lit,
                "Kelly 2013: 171",
            ],
            [
                "Hunter.max_size_water",
                "Per-group forager capacity, near water",
                H.max_size_water,
                "—",
                lit,
                "Kelly 2013: 171",
            ],
            [
                "Hunter.is_complex",
                "Forager sedentism threshold",
                H.is_complex,
                "—",
                lit,
                "Kelly 2013: 171",
            ],
            [
                "env.lim_h",
                "Forager capacity per cell",
                env.lim_h,
                ", ".join(str(v) for v in limh_swept()),
                lit,
                "Binford 2001; Tallavaara et al. 2017",
            ],
            [
                "Farmer.loss.prob / .rate",
                "Rainfed mortality, drawn once per step",
                f"{F.loss.prob} / {F.loss.rate}",
                "—",
                exp,
                "—",
            ],
            [
                "RiceFarmer.loss.prob / .rate",
                "Paddy mortality, drawn once per step",
                f"{R.loss.prob} / {R.loss.rate}",
                "—",
                exp,
                "—",
            ],
            [
                "Hunter.loss.prob / .rate",
                "Forager mortality, drawn once per step",
                f"{H.loss.prob} / {H.loss.rate}",
                "—",
                exp,
                "—",
            ],
            [
                "env.init_hunters",
                "Initial forager cover",
                env.init_hunters,
                "—",
                exp,
                "—",
            ],
            [
                "env.init_farmers",
                "Initial rainfed groups",
                env.init_farmers,
                "—",
                exp,
                "—",
            ],
            [
                "env.init_rice_farmers",
                "Initial paddy groups",
                env.init_rice_farmers,
                "—",
                exp,
                "—",
            ],
        ],
    )

    # ---- _index -----------------------------------------------------------
    idx = wb.create_sheet("_index")
    idx.append(["sheet", "label", "caption", "source", "updated"])
    for row in [
        [
            "Table 1",
            "tbl:key-params",
            "Baseline values and swept ranges for the parameters that govern how fast farming spreads.",
            "config/config.yaml; run_slurm_rerun.sh:150-151; data/*.tif via paper/build_tables.py",
            BUILT,
        ],
        [
            "Table S1",
            "tbl:stochasticity",
            "Every source of randomness in the model, with the reason each is random.",
            "paper/model-inventory.md section 5 (each entry carries a src file:line)",
            BUILT,
        ],
        [
            "Table S2",
            "tbl:conversion-paths",
            "The five directed conversion paths, with the conditions and probability gating each.",
            "src/api/hunter.py, farmer.py, rice_farmer.py; config/config.yaml; run_slurm_rerun.sh:150-151",
            BUILT,
        ],
        [
            "Table S3",
            "tbl:symbol-map",
            "One-to-one mapping between main-text symbols and implementation parameter names.",
            "paper/manuscript/methods.md; config/config.yaml",
            BUILT,
        ],
        [
            "Table S4",
            "tbl:parameters",
            "Every model parameter with baseline value, swept range, and provenance.",
            "config/config.yaml; run_slurm_rerun.sh:150-151; provenance per paper/model-inventory.md section 4",
            BUILT,
        ],
    ]:
        idx.append(row)
    for i, w in enumerate([12, 22, 70, 62, 12], start=1):
        idx.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(f"  landscape: {land}")
    print(f"  ceiling  : {ceiling}")
    print(f"  swept    : {swept}")


if __name__ == "__main__":
    build()
